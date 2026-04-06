"""
TaskFlow Scheduler Engine
─────────────────────────
Uses NetworkX for dependency graph analysis and topological sorting.
Assigns tasks to team members based on skill areas, availability, and
workload constraints.  Produces a Gantt-chart-ready schedule.
"""

from __future__ import annotations
import yaml, copy, datetime as dt
from dataclasses import dataclass, field
from typing import Optional
import networkx as nx


# ───────────────────────────── Data Models ────────────────────────────────

@dataclass
class Member:
    name: str
    areas: list[str]
    unavailable: list[tuple[dt.date, dt.date]] = field(default_factory=list)

    def is_available(self, day: dt.date) -> bool:
        return not any(s <= day <= e for s, e in self.unavailable)


@dataclass
class Task:
    id: int
    description: str
    areas: list[str]
    dependencies: list[int]
    estimated_days: int
    assigned_to: Optional[str] = None
    start_date: Optional[dt.date] = None
    end_date: Optional[dt.date] = None


@dataclass
class ScheduleResult:
    tasks: list[Task]
    graph: nx.DiGraph
    topo_order: list[int]
    warnings: list[str]
    members: list[Member]
    project_start: dt.date
    project_end: dt.date


# ───────────────────────────── YAML Parsing ───────────────────────────────

def _parse_date(s: str) -> dt.date:
    """Parse MM/DD/YYYY, YYYY-MM-DD, or MM-DD-YYYY date string."""
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"):
        try:
            return dt.datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Cannot parse date: {s!r}")


def _parse_date_ranges(raw_ranges) -> list[tuple[dt.date, dt.date]]:
    """Parse various date-range formats from YAML.

    Accepts:
      - [start, end] list/tuple pairs
      - "start, end" comma-separated strings
      - "start" single-date strings (treated as a one-day range)

    Malformed items are skipped rather than crashing.
    """
    if not raw_ranges:
        return []
    ranges = []
    for item in raw_ranges:
        try:
            if isinstance(item, (list, tuple)):
                if len(item) == 2:
                    ranges.append((_parse_date(str(item[0])), _parse_date(str(item[1]))))
                # length != 2: silently skip
            elif isinstance(item, str):
                parts = [p.strip() for p in item.strip("() ").split(",")]
                if len(parts) == 2:
                    ranges.append((_parse_date(parts[0]), _parse_date(parts[1])))
                elif len(parts) == 1 and parts[0]:
                    d = _parse_date(parts[0])
                    ranges.append((d, d))
        except ValueError:
            pass  # malformed date — skip
    return ranges


def load_yaml(path: str) -> dict:
    """Load and validate a project YAML file."""
    with open(path, "r") as f:
        raw = yaml.safe_load(f)

    project_start = _parse_date(str(raw["start_date"]))
    project_end   = _parse_date(str(raw["end_date"]))
    # Clamp to at least 1 so the scheduler never deadlocks.
    max_tasks = max(1, int(raw.get("maximum_tasks", 2)))

    members = []
    for m in raw.get("members", []):
        members.append(Member(
            name=m["name"],
            areas=[a.strip() for a in m.get("areas", [])],
            unavailable=_parse_date_ranges(m.get("dates_unavailable", [])),
        ))

    tasks = []
    seen_ids: set[int] = set()
    for raw_task in raw.get("tasks", []):
        task_id = raw_task.get("ID")
        if task_id is None:
            raise ValueError(f"A task is missing the required 'ID' field: {raw_task}")
        task_id = int(task_id)
        if task_id in seen_ids:
            raise ValueError(f"Duplicate task ID {task_id} found in YAML.")
        seen_ids.add(task_id)
        tasks.append(Task(
            id=task_id,
            description=raw_task.get("description", ""),
            areas=[a.strip() for a in raw_task.get("areas", [])],
            dependencies=[int(d) for d in raw_task.get("dependencies", [])],
            estimated_days=max(1, int(raw_task.get("estimated_time_days", 1))),
        ))

    return {
        "start":    project_start,
        "end":      project_end,
        "max_tasks": max_tasks,
        "members":  members,
        "tasks":    tasks,
    }


# ──────────────────────── Dependency Graph ────────────────────────────────

def build_graph(tasks: list[Task]) -> tuple[nx.DiGraph, list[str]]:
    """Build a directed dependency graph and return it with any warnings.

    Edge (A → B) means A must finish before B can start.
    Unknown dependency IDs are skipped with a warning rather than silently
    dropped.
    """
    G = nx.DiGraph()
    ids = {t.id for t in tasks}
    warnings: list[str] = []
    for t in tasks:
        G.add_node(t.id, task=t)
    for t in tasks:
        for dep in t.dependencies:
            if dep in ids:
                G.add_edge(dep, t.id)
            else:
                warnings.append(
                    f"Task {t.id}: dependency {dep} does not exist — edge skipped."
                )
    return G, warnings


def topological_order(G: nx.DiGraph) -> list[int]:
    """Return a topological ordering of the graph.

    Raises ValueError if the graph contains cycles.  Uses a single
    traversal (no redundant DAG pre-check).
    """
    try:
        return list(nx.topological_sort(G))
    except nx.NetworkXUnfeasible:
        cycles = list(nx.simple_cycles(G))
        cycle_str = ", ".join(str(c) for c in cycles[:5])
        raise ValueError(
            f"Circular dependencies detected — cannot schedule.\n"
            f"Cycles: {cycle_str}"
        )


# ──────────────────────── Scheduling Engine ───────────────────────────────

def _next_workday(day: dt.date) -> dt.date:
    """Advance to the next weekday (no-op if already a weekday)."""
    while day.weekday() >= 5:
        day += dt.timedelta(days=1)
    return day


def _advance_workdays(start: dt.date, n: int, member: Optional[Member] = None) -> dt.date:
    """Return the date that is *n* working days after (and including) *start*.

    If *member* is provided, days in the member's unavailability windows are
    not counted as working days, so the end date is correctly extended past
    any vacation periods.
    """
    current   = start
    remaining = n - 1          # start day counts as day 1
    while remaining > 0:
        current += dt.timedelta(days=1)
        if current.weekday() >= 5:
            continue           # weekend
        if member is not None and not member.is_available(current):
            continue           # member unavailable
        remaining -= 1
    return current


def schedule(data: dict) -> ScheduleResult:
    """Core scheduling algorithm.

    1. Build the dependency graph (topological sort; emit unknown-dep warnings).
    2. Walk tasks in topological order.  For each task:
       a. Skill quality — exact area matches preferred over partial matches;
          fall back to any member only if no skill match exists.
       b. Earliest available start — day-by-day scan respecting concurrency
          cap, weekends, and member unavailability.
       c. Workload balance — among equally-early candidates, prefer the member
          with the fewest tasks assigned so far.
    3. Compute end dates honouring the member's unavailability windows.
    4. Emit warnings for late tasks, skill mismatches, and broken deps.
    """
    tasks_by_id: dict[int, Task] = {t.id: copy.deepcopy(t) for t in data["tasks"]}
    members:     list[Member]    = copy.deepcopy(data["members"])
    max_concurrent = data["max_tasks"]
    project_start  = data["start"]
    project_end    = data["end"]

    G, warnings = build_graph(list(tasks_by_id.values()))
    topo = topological_order(G)

    # All assigned intervals per member — ground truth for concurrency checks.
    member_intervals:  dict[str, list[tuple[dt.date, dt.date]]] = {m.name: [] for m in members}
    # Total tasks assigned per member — tiebreaker for load balancing.
    member_task_count: dict[str, int] = {m.name: 0 for m in members}

    def _concurrent_on_day(name: str, day: dt.date) -> int:
        """Count overlapping assigned intervals on *day*."""
        return sum(1 for s, e in member_intervals[name] if s <= day <= e)

    def _earliest_start(member: Member, not_before: dt.date) -> dt.date:
        """First workday >= *not_before* where the member is below the
        concurrency cap and not in an unavailability window."""
        day = not_before
        for _ in range(730):
            # Normalise to weekday first, then check conditions.
            if day.weekday() >= 5:
                day += dt.timedelta(days=1)
                continue
            if not member.is_available(day):
                day += dt.timedelta(days=1)
                continue
            if _concurrent_on_day(member.name, day) < max_concurrent:
                return day
            day += dt.timedelta(days=1)
        return _next_workday(day)   # failsafe — always return a weekday

    for task_id in topo:
        task = tasks_by_id[task_id]

        # Earliest possible start from predecessor end dates.
        not_before = project_start
        for pred in G.predecessors(task_id):
            pred_task = tasks_by_id[pred]
            if pred_task.end_date:
                dep_end    = _next_workday(pred_task.end_date + dt.timedelta(days=1))
                not_before = max(not_before, dep_end)
            else:
                warnings.append(
                    f"Task {task_id}: predecessor {pred} has no end date "
                    f"(unassigned?) — dependency constraint ignored."
                )

        # Partition members by skill quality.
        exact:   list[Member] = []
        partial: list[Member] = []
        for m in members:
            if not task.areas:
                exact.append(m)
            elif all(a in m.areas for a in task.areas):
                exact.append(m)
            elif any(a in m.areas for a in task.areas):
                partial.append(m)

        skill_pool = exact or partial
        if not skill_pool:
            warnings.append(
                f"Task {task.id} ('{task.description}'): no member matches "
                f"areas {task.areas} — assigned to least-loaded available member."
            )
            skill_pool = members

        if not skill_pool:
            warnings.append(f"Task {task.id}: no members defined — skipped.")
            continue

        # Sort by (earliest start, fewest tasks) for workload balance.
        candidates = [
            (_earliest_start(m, not_before), member_task_count[m.name], m)
            for m in skill_pool
        ]
        candidates.sort(key=lambda x: (x[0], x[1]))

        start_date, _, chosen = candidates[0]
        end_date = _advance_workdays(start_date, task.estimated_days, chosen)

        if end_date > project_end:
            warnings.append(
                f"Task {task.id} ('{task.description}') ends {end_date}, "
                f"past the project deadline {project_end}."
            )

        task.assigned_to = chosen.name
        task.start_date  = start_date
        task.end_date    = end_date

        member_intervals[chosen.name].append((start_date, end_date))
        member_task_count[chosen.name] += 1

    return ScheduleResult(
        tasks=list(tasks_by_id.values()),
        graph=G,
        topo_order=topo,
        warnings=warnings,
        members=members,
        project_start=project_start,
        project_end=project_end,
    )


# ─────────────────────── Excel Gantt Export ────────────────────────────────

def export_gantt_xlsx(result: ScheduleResult, path: str):
    """Export schedule as a colour-coded Gantt chart in Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Task Schedule Table ──────────────────────────────────────
    ws = wb.active
    ws.title = "Schedule"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2B3A67")
    cell_font   = Font(name="Arial", size=10)
    warn_fill   = PatternFill("solid", fgColor="FFF3CD")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    headers = [
        "Task ID", "Description", "Areas", "Dependencies",
        "Assigned To", "Start Date", "End Date", "Duration (days)",
        "Topo Order"
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = header_font
        c.fill      = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border

    topo_pos    = {tid: i + 1 for i, tid in enumerate(result.topo_order)}
    sorted_tasks = sorted(result.tasks, key=lambda t: (t.start_date or dt.date.max))

    for row_idx, task in enumerate(sorted_tasks, 2):
        vals = [
            task.id,
            task.description,
            ", ".join(task.areas),
            ", ".join(str(d) for d in task.dependencies) if task.dependencies else "None",
            task.assigned_to or "Unassigned",
            task.start_date.strftime("%m/%d/%Y") if task.start_date else "",
            task.end_date.strftime("%m/%d/%Y")   if task.end_date   else "",
            task.estimated_days,
            topo_pos.get(task.id, ""),
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=v)
            c.font      = cell_font
            c.border    = thin_border
            c.alignment = Alignment(
                horizontal="center" if col != 2 else "left",
                vertical="center",
            )

    col_widths = [10, 35, 18, 16, 14, 14, 14, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Visual Gantt Chart ───────────────────────────────────────
    ws2 = wb.create_sheet("Gantt Chart")

    scheduled = [t for t in sorted_tasks if t.start_date and t.end_date]
    if not scheduled:
        wb.save(path)
        return

    min_date   = min(t.start_date for t in scheduled)
    max_date   = max(t.end_date   for t in scheduled)
    total_days = (max_date - min_date).days + 1

    member_colors: dict[str, str] = {}
    palette = ["4E79A7", "F28E2B", "E15759", "76B7B2", "59A14F",
               "EDC948", "B07AA1", "FF9DA7", "9C755F", "BAB0AC"]
    for i, m in enumerate(result.members):
        member_colors[m.name] = palette[i % len(palette)]

    weekend_shade = PatternFill("solid", fgColor="1A2340")
    row_weekend   = PatternFill("solid", fgColor="F0F0F0")

    # Precompute which spreadsheet columns are weekends (O(total_days) once).
    weekend_cols: set[int] = set()
    for d in range(total_days):
        if (min_date + dt.timedelta(days=d)).weekday() >= 5:
            weekend_cols.add(d + 3)

    # Headers row
    ws2.cell(row=1, column=1, value="Task").font    = header_font
    ws2.cell(row=1, column=1).fill                  = header_fill
    ws2.cell(row=1, column=1).border                = thin_border
    ws2.cell(row=1, column=2, value="Assignee").font = header_font
    ws2.cell(row=1, column=2).fill                  = header_fill
    ws2.cell(row=1, column=2).border                = thin_border
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 12

    for d in range(total_days):
        day = min_date + dt.timedelta(days=d)
        col = d + 3
        c          = ws2.cell(row=1, column=col)
        c.value    = day.strftime("%m/%d")
        c.font     = Font(name="Arial", size=7, bold=True, color="FFFFFF")
        c.fill     = weekend_shade if col in weekend_cols else header_fill
        c.alignment = Alignment(horizontal="center", text_rotation=90)
        c.border   = thin_border
        ws2.column_dimensions[get_column_letter(col)].width = 3.5

    # Task bars — track filled cells to avoid overwriting bars with weekend shade.
    bar_filled: set[tuple[int, int]] = set()

    for row_idx, task in enumerate(sorted_tasks, 2):
        ws2.cell(row=row_idx, column=1,
                 value=f"[{task.id}] {task.description}").font = cell_font
        ws2.cell(row=row_idx, column=1).border = thin_border
        ws2.cell(row=row_idx, column=2,
                 value=task.assigned_to or "").font = cell_font
        ws2.cell(row=row_idx, column=2).border = thin_border

        if task.start_date and task.end_date:
            color    = member_colors.get(task.assigned_to, "999999")
            bar_fill = PatternFill("solid", fgColor=color)
            start_col = (task.start_date - min_date).days + 3
            end_col   = (task.end_date   - min_date).days + 3
            for col in range(start_col, end_col + 1):
                c        = ws2.cell(row=row_idx, column=col)
                c.fill   = bar_fill
                c.border = thin_border
                bar_filled.add((row_idx, col))

        # Apply weekend shading only to unfilled cells.
        for col in weekend_cols:
            if (row_idx, col) not in bar_filled:
                c        = ws2.cell(row=row_idx, column=col)
                c.fill   = row_weekend
                c.border = thin_border

    # ── Sheet 3: Dependency Info ──────────────────────────────────────────
    ws3 = wb.create_sheet("Dependencies")
    dep_headers = ["Task ID", "Description", "Depends On (IDs)",
                   "Depended By (IDs)", "Topological Position"]
    for col, h in enumerate(dep_headers, 1):
        c        = ws3.cell(row=1, column=col, value=h)
        c.font   = header_font
        c.fill   = header_fill
        c.border = thin_border

    task_by_id = {t.id: t for t in result.tasks}   # O(1) lookups below
    for row_idx, tid in enumerate(result.topo_order, 2):
        task  = task_by_id[tid]
        preds = list(result.graph.predecessors(tid))
        succs = list(result.graph.successors(tid))
        vals  = [
            tid,
            task.description,
            ", ".join(str(p) for p in preds) if preds else "None",
            ", ".join(str(s) for s in succs) if succs else "None",
            row_idx - 1,
        ]
        for col, v in enumerate(vals, 1):
            c        = ws3.cell(row=row_idx, column=col, value=v)
            c.font   = cell_font
            c.border = thin_border

    for i, w in enumerate([10, 35, 20, 20, 18], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 4: Warnings ─────────────────────────────────────────────────
    if result.warnings:
        ws4 = wb.create_sheet("Warnings")
        ws4.cell(row=1, column=1, value="Warning").font = header_font
        ws4.cell(row=1, column=1).fill = PatternFill("solid", fgColor="CC0000")
        ws4.column_dimensions["A"].width = 80
        for i, w in enumerate(result.warnings, 2):
            c      = ws4.cell(row=i, column=1, value=w)
            c.font = cell_font
            c.fill = warn_fill

    wb.save(path)


def export_gantt_csv(result: ScheduleResult, path: str):
    """Export a simple CSV of the schedule."""
    import csv
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Task ID", "Description", "Areas", "Dependencies",
            "Assigned To", "Start Date", "End Date", "Duration (days)"
        ])
        for t in sorted(result.tasks, key=lambda t: (t.start_date or dt.date.max)):
            writer.writerow([
                t.id,
                t.description,
                "; ".join(t.areas),
                "; ".join(str(d) for d in t.dependencies),
                t.assigned_to or "Unassigned",
                t.start_date.strftime("%m/%d/%Y") if t.start_date else "",
                t.end_date.strftime("%m/%d/%Y")   if t.end_date   else "",
                t.estimated_days,
            ])


# ─────────────────────── ICS Calendar Export ──────────────────────────────

def _ics_escape(text: str) -> str:
    """Escape special characters for iCalendar text fields."""
    return (text
            .replace("\\", "\\\\")
            .replace(";",  "\\;")
            .replace(",",  "\\,")
            .replace("\n", "\\n"))


def _fold_line(line: str) -> str:
    """Fold long lines per RFC 5545 (max 75 octets per line)."""
    if len(line.encode("utf-8")) <= 75:
        return line
    result  = []
    current = ""
    for ch in line:
        test = current + ch
        if len(test.encode("utf-8")) > 75:
            result.append(current)
            current = " " + ch      # continuation line starts with LWSP
        else:
            current = test
    if current:
        result.append(current)
    return "\r\n".join(result)


def export_ics(result: ScheduleResult, path: str, alarm_minutes: int = 30):
    """Export schedule as an ICS (iCalendar) file.

    Each task becomes an all-day VEVENT spanning from start_date to
    end_date + 1 day (exclusive per iCal spec for DATE values).
    """
    import uuid

    now_utc = dt.datetime.now(dt.timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    prod_id = "-//TaskFlow//Project Scheduler//EN"

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        f"PRODID:{prod_id}",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        f"X-WR-CALNAME:TaskFlow – {result.project_start.strftime('%b %Y')}",
        "X-WR-TIMEZONE:UTC",
    ]

    palette_names = ["Blue", "Orange", "Red", "Teal", "Green",
                     "Yellow", "Purple", "Pink", "Brown", "Gray"]
    member_colors = {
        m.name: palette_names[i % len(palette_names)]
        for i, m in enumerate(result.members)
    }
    # O(1) position lookup instead of O(N) list.index() per task.
    topo_pos = {tid: i + 1 for i, tid in enumerate(result.topo_order)}

    sorted_tasks = sorted(result.tasks, key=lambda t: (t.start_date or dt.date.max))

    for task in sorted_tasks:
        if not task.start_date or not task.end_date:
            continue

        uid     = f"taskflow-{task.id}-{uuid.uuid4().hex[:8]}@taskflow"
        dtstart = task.start_date.strftime("%Y%m%d")
        dtend   = (task.end_date + dt.timedelta(days=1)).strftime("%Y%m%d")

        preds = list(result.graph.predecessors(task.id))
        succs = list(result.graph.successors(task.id))
        desc_parts = [
            f"Task ID: {task.id}",
            f"Areas: {', '.join(task.areas) if task.areas else 'General'}",
            f"Assigned to: {task.assigned_to or 'Unassigned'}",
            f"Duration: {task.estimated_days} working days",
            "",
            f"Dependencies (upstream): {', '.join(f'Task {p}' for p in preds) if preds else 'None'}",
            f"Feeds into (downstream): {', '.join(f'Task {s}' for s in succs) if succs else 'None'}",
            "",
            f"Topological position: {topo_pos.get(task.id, '?')} of {len(result.topo_order)}",
        ]
        # Join with real newlines so _ics_escape converts them to \\n exactly once.
        description = _ics_escape("\n".join(desc_parts))
        summary     = _ics_escape(f"[Task {task.id}] {task.description}")
        category    = member_colors.get(task.assigned_to, "Blue")

        lines.extend([
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTAMP:{now_utc}",
            f"DTSTART;VALUE=DATE:{dtstart}",
            f"DTEND;VALUE=DATE:{dtend}",
            f"SUMMARY:{summary}",
            f"DESCRIPTION:{description}",
            f"CATEGORIES:{category}",
            "STATUS:CONFIRMED",
            "TRANSP:TRANSPARENT",
        ])

        if task.assigned_to:
            safe_name = _ics_escape(task.assigned_to)
            lines.append(
                f"ATTENDEE;CN={safe_name};ROLE=REQ-PARTICIPANT:"
                f"mailto:{task.assigned_to.lower().replace(' ', '.')}@taskflow.local"
            )

        if alarm_minutes and alarm_minutes > 0:
            lines.extend([
                "BEGIN:VALARM",
                f"TRIGGER:-PT{alarm_minutes}M",
                "ACTION:DISPLAY",
                f"DESCRIPTION:Task starting: {_ics_escape(task.description)}",
                "END:VALARM",
            ])

        lines.append("END:VEVENT")

    lines.append("END:VCALENDAR")

    with open(path, "w", encoding="utf-8", newline="") as f:
        for line in lines:
            f.write(_fold_line(line) + "\r\n")


def generate_calendar_links(result: ScheduleResult) -> list[dict]:
    """Generate Google Calendar and Outlook Web 'Add Event' URLs for each task."""
    import urllib.parse

    # O(1) position lookup.
    topo_pos = {tid: i + 1 for i, tid in enumerate(result.topo_order)}

    links = []
    for task in sorted(result.tasks, key=lambda t: (t.start_date or dt.date.max)):
        if not task.start_date or not task.end_date:
            continue

        title = f"[Task {task.id}] {task.description}"
        preds = list(result.graph.predecessors(task.id))
        succs = list(result.graph.successors(task.id))
        details = (
            f"Task ID: {task.id}\n"
            f"Areas: {', '.join(task.areas)}\n"
            f"Assigned to: {task.assigned_to or 'Unassigned'}\n"
            f"Duration: {task.estimated_days} working days\n"
            f"Depends on: {', '.join(f'Task {p}' for p in preds) if preds else 'None'}\n"
            f"Feeds into: {', '.join(f'Task {s}' for s in succs) if succs else 'None'}\n"
            f"Topological position: {topo_pos.get(task.id, '?')} of {len(result.topo_order)}"
        )

        dtstart = task.start_date.strftime("%Y%m%d")
        dtend   = (task.end_date + dt.timedelta(days=1)).strftime("%Y%m%d")

        google_params = urllib.parse.urlencode({
            "action":  "TEMPLATE",
            "text":    title,
            "dates":   f"{dtstart}/{dtend}",
            "details": details,
            "ctz":     "UTC",
        })
        google_url = f"https://calendar.google.com/calendar/render?{google_params}"

        outlook_params = urllib.parse.urlencode({
            "rru":      "addevent",
            "startdt":  task.start_date.isoformat(),
            "enddt":    (task.end_date + dt.timedelta(days=1)).isoformat(),
            "subject":  title,
            "body":     details,
            "allday":   "true",
        })
        outlook_url = f"https://outlook.live.com/calendar/0/action/compose?{outlook_params}"

        links.append({
            "task_id":     task.id,
            "description": task.description,
            "assigned_to": task.assigned_to,
            "start":       task.start_date,
            "end":         task.end_date,
            "google_url":  google_url,
            "outlook_url": outlook_url,
        })

    return links
